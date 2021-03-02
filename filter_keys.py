import openpyxl
from openpyxl.styles import Font

# see examples on https://linuxhint.com/manipulating_excel_python/
# Initializing a Workbook
work_book = openpyxl.load_workbook('words_with_dublicate.xlsx')

print(work_book.sheetnames)

# Navigating to Second Sheet (at index 1)
work_book.active = 0
# Getting Active Sheet
sheet = work_book.active
print(sheet.title)
# 1 Удаляем все дубликаты - проходим по ключам с конца, и перезаписываем повторение.
sheet_values = tuple(list(sheet.values)[::-1])
result = {sheet_values[i][0]: (sheet_values[i][1], sheet_values[i][2])
          for i in range(len(sheet_values))}
print(result)
# Инициализация новой книги
work_book = openpyxl.Workbook()
work_book.create_sheet(index=0, title='Words')
work_book.active = 0
# Getting Active Sheet
sheet = work_book.active
for n, (key, values) in enumerate(result.items(), 1):
    sheet[f'A{n}'] = key
    sheet[f'B{n}'] = values[0]
    sheet[f'C{n}'] = values[1]

work_book.save('words_for_direct.xlsx')